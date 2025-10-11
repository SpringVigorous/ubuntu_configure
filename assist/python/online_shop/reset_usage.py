import os

import sys
from openpyxl import load_workbook


from base import logger_helper,UpdateTimeType,unmerge_sheet_cols,read_xlsx,delete_sheets_except,unique,get_col_index
from openpyxl import load_workbook


#谨慎使用，若是dest_path为None，则覆盖原文件
def reset_usage(xlsx_path,unmerge_key_cols:list|str,key_col_name:str,sheet_name:str=None,unmerge_col_names:list[str]=None,dest_path:str=None,dest_sheet_name:str=None) -> None:
    """重置用量表
    Args:
        xlsx_path (str): xlsx文件路径
        sheet_name (str, optional): 工作表名称. Defaults to None.
        col_index (list[int], optional): 需要重置的列索引. Defaults to None.
    """

    ws=read_xlsx(xlsx_path,sheet_name)
    if not ws:
        return
    wb=ws.parent
    
    if dest_sheet_name:
        ws.title=dest_sheet_name
    unmerge_sheet_cols(ws,unmerge_key_cols)
    delete_sheets_except(wb,ws.title)
    if not dest_path:
        dest_path=xlsx_path
        

    column_num = get_col_index(ws,key_col_name)
    if column_num:
        # 收集有效值(非空且非空字符串)
        valid_values = []
        for cell in ws.iter_cols(min_col=column_num, max_col=column_num):
            for c in cell:
                if c.value is not None and str(c.value).strip() != "":
                    valid_values.append(str(c.value).strip())
                    
        valid_values=unique(valid_values)
        if valid_values:
            new_sheet_name="产品名"
            # 创建新工作表(如果已存在则删除旧版)
            if new_sheet_name in wb.sheetnames:
                del wb[new_sheet_name]
            new_sheet = wb.create_sheet(new_sheet_name)
            
            # 将值写入新工作表的第一列
            for row, value in enumerate(valid_values, start=1):
                new_sheet.cell(row=row, column=1, value=value)
        
        
    wb.save(dest_path)
    
    
    
if __name__=="__main__":
    key_cols=['产品','药材','实际规格']
    key_col_name="产品"
    reset_usage(r"E:\花茶\花茶定制.xlsx",key_cols,key_col_name,dest_path=r"E:\花茶\价格\产品用量.xlsx",sheet_name="用量",dest_sheet_name="用量")