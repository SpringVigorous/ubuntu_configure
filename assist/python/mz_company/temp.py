import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# 创建 DataFrame
df_stock = pd.DataFrame({
    '商品编码': ['000001', '000001', '000001', '000001', '000002'],
    '商品名称': ['商品1', '商品2', '商品1', '商品2', '商品3'],
    '价格': [100, 100, 200, 100, 300],
    '成本价': [10.0, 10.0, 20.0, 20.0, 30.0],
    "id": [1, 1, 2, 3, 3]
})

# 创建一个新的Excel工作簿
wb = openpyxl.Workbook()
ws = wb.active


# 设置表头
headers = df_stock.columns.tolist()
for col_idx, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col_idx, value=header)
    ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')

# 将 DataFrame 写入 Excel 工作表
for r in dataframe_to_rows(df_stock, index=False, header=False):
    ws.append(r)

# 自动合并单元格
def merge_identical_column_cells(ws, col_idx, start_row, end_row):
    current_value = ws.cell(row=start_row, column=col_idx).value
    merge_start = start_row
    
    for row in range(start_row + 1, end_row + 1):
        cell_value = ws.cell(row=row, column=col_idx).value
        if cell_value == current_value:
            continue
        else:
            if merge_start != row - 1:
                ws.merge_cells(start_row=merge_start, start_column=col_idx, end_row=row - 1, end_column=col_idx)
                ws.cell(row=merge_start, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
            current_value = cell_value
            merge_start = row
    
    # 合并最后一组相同的单元格
    if merge_start != end_row:
        ws.merge_cells(start_row=merge_start, start_column=col_idx, end_row=end_row, end_column=col_idx)
        ws.cell(row=merge_start, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')

# 获取数据行数
end_row = ws.max_row

# 合并每一列的相同单元格
for col_idx in range(1, len(headers) + 1):
    merge_identical_column_cells(ws, col_idx, 2, end_row)

# 设置列宽
ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 20
ws.column_dimensions["E"].width = 20

# 保存Excel文件
excel_path = 'grouped_stock.xlsx'
wb.save(excel_path)

print(f"分组结果已保存到 {excel_path}")