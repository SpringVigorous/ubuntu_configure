import pandas as pd
import os
import requests
from PIL import Image as PILImage
import re
import aiohttp
import asyncio

import sys

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment


from base.xls_tools import *
from base.file_tools import download_async

from pathlib import Path

"""
"季节",
"主图",
"商家编码",
"颜色",
"100",
"110",
"120",
"总库存",
"成本价",
"总金额",

"""
#原始表格数据
def get_org_df(org_path:Path,sheet_name):
    org_df=pd.read_excel(org_path,sheet_name=sheet_name, dtype=str)
    org_df.dropna(subset=['货号'],inplace=True)
    org_df.loc[:,'成本价']=org_df['成本价'].astype(float).fillna(0)
    org_df.loc[:,'库存']=org_df['库存'].fillna(0).astype(int)
    

    org_df.sort_values(by=['商家编码',"颜色","尺码"],ascending=[True,True,True],inplace=True)
    temp_path=org_path.parent.joinpath(f"{org_path.stem}_临时{org_path.suffix}")
    org_df.to_excel(temp_path,sheet_name=sheet_name,index=False)
    return org_df


goods_headers=["季节",
    "主图",
    "商家编码"]

stock_headers=["商品编码","商家编码","颜色","尺码","库存","成本价"]


#获取图片路径，不存在则为空
def get_picture_path(picture_dir,code):
    picture_path=os.path.join(picture_dir,f"{code}.jpg")
    if not os.path.exists(picture_path):
        picture_path=""
    return picture_path
#商品信息
def get_goods_df(org_df,org:Path):
    goods_df=org_df[goods_headers].copy()
    #根据 商家编码 去重
    goods_df.drop_duplicates(subset=['商家编码'],inplace=True)
    pictrue_dir=org.parent.joinpath("图片")
    goods_df["主图"]=goods_df["商家编码"].apply(lambda x: get_picture_path(pictrue_dir,x))
    
    temp_path=org.parent.joinpath("临时",f"{org.stem}_商品信息{org.suffix}")
    os.makedirs(temp_path.parent,exist_ok=True)
    goods_df.to_excel(temp_path,sheet_name="Sheet1",index=False)
    return goods_df
#库存信息
def get_stock_df(org_df,org:Path):
    stock_df=org_df[stock_headers].copy()
    temp_path=org.parent.joinpath("临时",f"{org.stem}_库存信息{org.suffix}")
    os.makedirs(temp_path.parent,exist_ok=True)
    stock_df.to_excel(temp_path,sheet_name="Sheet1",index=False)
    return stock_df


#合并信息
def get_dest_df(goods_df,stock_df,size_lst,org:Path):
    merged_df = pd.merge(goods_df, stock_df, on='商家编码', how='inner').drop(columns=["商品编码"])
    # 按照 '商家编码' 列进行第一次分组
    grouped_by_merchant = merged_df.groupby("商家编码")
    headers=["季节",
    "主图",
    "商家编码",
    "颜色",]
    
    headers.append("颜色")   
    headers.extend(size_lst)
    headers.extend(["总库存","成本价","总金额"])
    dest_df=pd.DataFrame( {item:[] for item in headers}  )
    #写入头文件

    for code,item in grouped_by_merchant:
        picture_path=item["主图"].iloc[0]
        season=item["季节"].iloc[0]
        
        color_grouped = item.groupby("颜色")
        for color,color_item in color_grouped:           
            stock=color_item["库存"].sum()
            price=color_item["成本价"].iloc[0]
            row_data={
                "季节":season,
                "主图":picture_path,
                "商家编码":code,
                "颜色":color,
                "总库存":stock,
                "成本价":price,
                "总金额":stock*price,
            }
            for index,row in color_item.iterrows():
                stock_val=row["库存"]
                if stock_val>0:
                    row_data[row["尺码"]]=stock_val
            dest_df.loc[len(dest_df)] = row_data
            
            
    if "0" in dest_df.columns:
        dest_df.rename(columns={"0":"无尺寸"},inplace=True)
        
    temp_path=org.parent.joinpath("临时",f"{org.stem}_mid{org.suffix}")
    os.makedirs(temp_path.parent,exist_ok=True)
    
    dest_df.to_excel(temp_path,index=False)

    return dest_df


def export_to_excel(mid_df,org:Path):
   
    wb = Workbook()
    ws = wb.active
    ws.title = "库存信息"

    write_dataframe_content(ws,mid_df,False,True)
    #统计数量
    count_df=  mid_df.groupby("商家编码",sort=False).agg({"颜色":"count"}).reset_index()
    count_map= count_df.set_index("商家编码")["颜色"].to_dict()
    
    #合并 主图、商家编码
    merge_identical_column_cells(ws,2)
    merge_identical_column_cells(ws,3)
    season_col_index=1
    show_height=80
    each_row_height= 13.5
    
    end_row_idx=1

    
    for code,count in count_map.items():
        start_row_idx=end_row_idx+1
        end_row_idx=start_row_idx+count-1
        #合并季节列
        if count>1:
            merge_cells_center(ws,start_row_idx,season_col_index,end_row_idx,season_col_index)
        
        #添加图片
        picture_cell=ws.cell(row=start_row_idx, column=2)
        if not picture_cell.value :
            continue
        add_image_to_cell(picture_cell,picture_cell.value)
        
        #设置行高
        cur_row_height=each_row_height
        total_height=each_row_height*count
        if total_height<show_height:
            cur_row_height=show_height/count
            
        for row_idx in range(start_row_idx,end_row_idx+1):
            ws.row_dimensions[row_idx].height =cur_row_height
            
        # ws.row_dimensions[cell.row].height =show_height
        picture_cell.value=""
        
        #居中对齐
        if count==1:
            set_cell_center(ws.cell(start_row_idx,1))
            set_cell_center(ws.cell(start_row_idx,3))
        
        pass
    



    ws.freeze_panes = 'A2'  # 冻结首行
    ws.column_dimensions["B"].width = 15
    
    
    dest_path=org.parent.joinpath("结果",f"{org.stem}_处理{org.suffix}")
    wb.save(dest_path)
    

def handle_xlsx(org_path,sheet_name):
    org=Path(org_path)
    org_df=get_org_df(org,sheet_name)
    size_lst= list(dict.fromkeys(org_df["尺码"].to_list()))

    goods_df=get_goods_df(org_df=org_df,org=org)
    stock_df=get_stock_df(org_df=org_df,org=org)
    dest_df= get_dest_df(goods_df,stock_df,size_lst,org)
    export_to_excel(dest_df,org)

if __name__=="__main__":
    handle_xlsx(r"E:\公司文件\库存\副本全店价格表1月2号.xlsx","Sheet1")