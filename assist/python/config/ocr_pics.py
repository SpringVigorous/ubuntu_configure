import pytesseract
from PIL import Image, ImageFilter
import openpyxl
import os
from pathlib import Path

def get_jpg_files(folder_path):
    folder = Path(folder_path)
    jpg_files = [str(file) for file in folder.rglob('*.jpg')]
    return jpg_files

# 获取 .jpg 文件路径集合
folder_path = r'F:\教程\短视频教程\抖音\QQ拉新粉\交流群成员'
image_paths = get_jpg_files(folder_path)




# 创建一个新的工作簿和工作表
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "文字识别结果"

# 写入表头
ws.cell(row=1, column=1, value="图片文件名")
ws.cell(row=1, column=2, value="识别结果")

# 从第二行开始写入数据
row = 2

for image_path in image_paths:
    # 打开图片
    img = Image.open(image_path)
    
    # 图像预处理
    img = img.convert('L')  # 转换为灰度图像
    img = img.filter(ImageFilter.MedianFilter())  # 中值滤波去噪
    
    # 使用 pytesseract 进行文字识别，指定语言为简体中文
    text = pytesseract.image_to_string(img, lang='chi_sim')
    
    # 将结果写入 Excel
    ws.cell(row=row, column=1, value=image_path)
    ws.cell(row=row, column=2, value=text)
    
    row += 1

# 保存 Excel 文件
wb.save('文字识别结果.xlsx')