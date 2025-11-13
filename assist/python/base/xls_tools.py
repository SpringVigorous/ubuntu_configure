from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from base.collect_tools import unique






def set_cell_center(cell:Cell):
    cell.alignment = Alignment(horizontal='center', vertical='center')

#合并单元格，并设置居中对齐
def merge_cells_center(ws:Worksheet, start_row, start_column, end_row, end_column,lat_func=None):
    ws.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)
    cell=ws.cell(row=start_row, column=start_column)
    set_cell_center(cell=cell)

    
    #后续操作
    lat_func(ws,cell.coordinate) if lat_func else None
    
#按列合并单元格
def merge_column_cells(ws, start_row, start_col, row_count=-1):
    merge_cells_center(ws, start_row=start_row, start_column=start_col, end_row=start_row+row_count-1 if row_count!=-1 else ws.max_row, end_column=start_col)

#仅写入表头，并设置居中对齐
def write_headers(ws:Worksheet, headers, start_row=1, start_column=1):
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col_idx, value=header)
        ws.cell(row=start_row, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
# 将 DataFrame 写入 Excel 工作表，可选择是否包含 表头
def write_dataframe_content(ws:Worksheet, df_stock:pd.DataFrame, index=False, header=False):
    for r in dataframe_to_rows(df_stock, index=index, header=header):
        ws.append(r)


#按列  自动合并单元格
def merge_identical_column_cells(ws:Worksheet, col_idx, start_row=2, end_row=-1,lat_func=None):
    if end_row==-1:
        end_row = ws.max_row
    
    current_value = ws.cell(row=start_row, column=col_idx).value
    merge_start = start_row
    
    for row in range(start_row + 1, end_row + 1):
        cell_value = ws.cell(row=row, column=col_idx).value
        if cell_value == current_value:
            continue
        else:
            if merge_start != row - 1:
                merge_cells_center(ws,start_row=merge_start, start_column=col_idx, end_row=row - 1, end_column=col_idx,lat_func=lat_func)
            current_value = cell_value
            merge_start = row
    
    # 合并最后一组相同的单元格
    if merge_start != end_row:
       merge_cells_center(ws,start_row=merge_start, start_column=col_idx, end_row=end_row, end_column=col_idx,lat_func=lat_func)

        
# 按列 自动合并相同行的单元格
def merge_all_identical_column_cells(ws:Worksheet,start_row=2, end_row=-1):
    for col_idx in range(1, ws.max_column + 1):
        merge_identical_column_cells(ws, col_idx, start_row, end_row)


#合并单元格的区域信息
def merge_cells_info(target_cell:Cell):
    if not target_cell:
        return None
    ws:Worksheet=target_cell.parent
    merged_ranges = list(ws.merged_cells.ranges)
    
    for merged_range in merged_ranges:
        # 获取合并单元格的左上角单元格的值
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_cell = ws.cell(row=min_row, column=min_col)
        top_left_value = top_left_cell.value
        
        # 检查 target_cell 是否在 merged_range 内
        if target_cell:
            target_row = target_cell.row
            target_col = target_cell.column
            if (min_row <= target_row <= max_row and min_col <= target_col <= max_col):
                return merged_range    
    return None


#合并单元格原始尺寸:获取合并单元格的范围
def merge_cells_original_size(target_cell:Cell):
    info =merge_cells_info(target_cell)
    return info.bounds if info else None


#单元格添加图片
def add_image_to_cell(cell:Cell, image_path,width=100,height=100):
    if not cell:
        return

    ws:Worksheet=cell.parent
    
    img = Image(image_path)
    # 调整图片大小（可选）

    
    src_width,src_height=img.width,img.height
    
    src_scale=float(src_width)/src_height
    if src_scale>float(width)/height:
        img.width = width
        img.height = int(width/src_scale)
    else:
        img.height = height
        img.width = int(height*src_scale)
    
    # img.width = width
    # img.height = height
    ws.add_image(img, cell.coordinate)

    
    
def unmerge_cells(ws:Worksheet, merged_range):
    # 获取合并单元格的左上角单元格的值
    min_col, min_row, max_col, max_row = merged_range.bounds
    top_left_cell = ws.cell(row=min_row, column=min_col)
    top_left_value = top_left_cell.value
    
    if top_left_value and isinstance(top_left_value,str):
        top_left_value= top_left_value.strip()
        top_left_cell.value=top_left_value
    # 取消合并单元格
    ws.unmerge_cells(str(merged_range))
    
    # 填充合并单元格内的所有单元格为左上角单元格的值
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = top_left_value
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
def unmerge_and_fill(target: Worksheet|Cell,filter_func:list[int,int,int,int]=None):
    if not target:
        return
    ws:Worksheet= target if isinstance(target, Worksheet) else target.parent
    target_cell:Cell=target if isinstance(target, Cell) else None
    #针对参数是 单元格类型的
    if target_cell:
        info=merge_cells_info(target_cell)
        if info:
            unmerge_cells(ws, info)
        return
    

    # 获取合并单元格的范围
    for merged_cell in list(ws.merged_cells.ranges):
        if filter_func and not filter_func(merged_cell.bounds):
            continue
        unmerge_cells(ws, merged_cell)

#针对首行值唯一的情况（及列名唯一）
def get_first_row_non_empty_cells(worksheet)->dict:
    """
    获取工作表首行中的非空单元格及其列号
    
    参数:
        worksheet: openpyxl.worksheet.worksheet.Worksheet 对象
    
    返回:
        list: 包含元组的列表，每个元组格式为 (列号, 单元格值)
    """
    first_row = worksheet[1]  # 获取第一行（索引从1开始）
    result = {}
    
    for cell in first_row:
        if cell.value is not None and str(cell.value).strip() != "":
            # 获取列字母（如'A'）和列号（如1）
            # column_letter = cell.column_letter
            column_number = cell.column
            result[cell.value]=column_number
    
    return result

def get_col_index(ws: Worksheet,col_name) -> int:
    col_dict=get_first_row_non_empty_cells(ws)

    col_index=col_dict.get(col_name,None)
    return col_index
    
    
#按指定列名取消 合并单元格
def unmerge_sheet_cols(ws: Worksheet,col_names:list|str=None):
    if not ws:
        return
    if not col_names:
        return unmerge_and_fill(ws,None)
       
    col_dict=get_first_row_non_empty_cells(ws)
    if isinstance(col_names,str):
        col_names=[col_names]
        
    cols_index=[col_dict.get(col_name,None) for col_name in unique(col_names)]
    cols_index=list(filter(lambda x:x,cols_index))
    
    for merged_cell in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merged_cell.bounds
        for i in cols_index:
            if i>=min_col and i<=max_col:
                unmerge_cells(ws, merged_cell)

  
def merge_all_identical_column_file(xlsx_path,sheet_name:str=None,col_index:list[int]=None):
    wb = load_workbook(xlsx_path)

    ws = wb.active if not sheet_name else wb[sheet_name]
    if not col_index:
        merge_all_identical_column_cells(ws)
    else:
        for col_idx in col_index:
            merge_identical_column_cells(ws, col_idx)
    # 保存工作簿
    wb.save(xlsx_path)
 

def read_xlsx(xlsx_path,sheet_name:str=None):
    wb=load_workbook(xlsx_path)
    if sheet_name and sheet_name in wb.sheetnames:
        ws=wb[sheet_name]
    else:
        ws=wb.active
    
    return ws

def delete_sheets_except(wb, sheet_name_to_keep):
    """
    删除工作簿中除指定名称表单外的所有其他表单
    
    参数:
        wb: openpyxl Workbook 对象
        sheet_name_to_keep: 要保留的表单名称
    """
    # 获取所有表单名称
    sheet_names = [sheet.title for sheet in wb.worksheets]
    
    # 检查要保留的表单是否存在
    if sheet_name_to_keep not in sheet_names:
        raise ValueError(f"表单 '{sheet_name_to_keep}' 不存在于工作簿中")
    
    # 遍历所有表单进行删除
    for sheet_name in sheet_names:
        if sheet_name != sheet_name_to_keep:
            # 删除表单
            del wb[sheet_name]
                
if __name__=="__main__":
    
    # 示例使用
    file_path=r'E:\公司文件\库存\结果\merge_result.xlsx'
    wb = load_workbook(file_path)
    ws = wb.active

    
    for col_idx in range(1, 6):
        merge_identical_column_cells(ws, col_idx)
    # merge_all_identical_column_cells(ws)
    
    # # 取消合并单元格 A1 并填充为左上角单元格的值
    # unmerge_and_fill(ws.cell(row=2, column=2))
    
    # def func(bounds):
    #     min_col, min_row, max_col, max_row = bounds
    #     return min_col>1
    
    # unmerge_and_fill(ws,func)
    
    # 保存工作簿
    wb.save(file_path)