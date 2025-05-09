import re
import openpyxl
from openpyxl import Workbook

def extract_table_statements(sql_content):
    """
    从 SQL 内容中提取所有 CREATE TABLE 语句
    """
    # 匹配 CREATE TABLE 语句（跨行匹配，直到分号结束）
    create_table_pattern = re.compile(
        r'CREATE\s+TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?(.*?\);?)', 
        re.DOTALL | re.IGNORECASE
    )
    return create_table_pattern.findall(sql_content)

def parse_table_statement(block):

    if not block : return None
    tab_name,other_str=block
    table_comment=""
    last_bracket_index = other_str.rfind(")")
    body_part=other_str[:last_bracket_index]
    comment_part=other_str[last_bracket_index+1:]
    
    tab_comment_partern=re.compile(r'COMMENT\s*=\s*[\'"](.*?)[\'"]', re.IGNORECASE)
    filed_comment_partern=re.compile(r'COMMENT\s*[\'"](.*?)[\'"]', re.IGNORECASE)
    field_partern=re.compile(r'`?(\w+)`?\s+(.*)', re.DOTALL | re.IGNORECASE)
    filed_sub_pattern = re.compile(
        r'\b(?:UNIQUE\s+KEY|FOREIGN\s+KEY|INDEX)\b\s*',  # 匹配目标关键词及空格
        flags=re.IGNORECASE  # 忽略大小写
    )

    
    if comment_part:
        # 提取表注释
        comment_match = tab_comment_partern.search(
            comment_part
        )
        if comment_match:
            table_comment = comment_match.group(1)
    fields=[]
    

    # 2. 解析字段定义段
    if body_part:
        fields_section = body_part
        # 清理字段段（处理多行注释）
        clean_fields = re.sub(r'--.*?(\n|$)', '', fields_section)
        # 安全分割字段（避免分割含括号内容）
        for field in re.split(r',\s*(?![^()]*\))', clean_fields):
            field = filed_sub_pattern.sub("",field).strip() 
            
            if not field:
                continue
            # 提取字段名和注释

            field_match = field_partern.search(field)
            if field_match:
                in_name,in_comment_part=field_match.groups()
                in_coment=""
                comment_match = filed_comment_partern.search(in_comment_part)
                if comment_match:
                    in_coment = comment_match.group(1)

        
                
                
                fields.append({
                    'name': in_name,
                    'comment': in_coment
                })

    table_info = {
        "table_name": tab_name,
        "fields": fields,
        "table_comment": table_comment
    }

    return table_info
def parse_sql_file(sql_file_path):
    """
    优化后的解析逻辑，适配您提供的 SQL 格式
    """
    tables = []
    content=None
    with open(sql_file_path, 'r', encoding='utf-8-sig') as f:
        content = f.read()
    if not content:
        return tables
    """主流程函数"""
    with open(sql_file_path, 'r', encoding='utf-8-sig') as f:
        sql_content = f.read()

    # 提取所有 CREATE TABLE 语句
    statements = re.findall(
        r'CREATE\s+TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?`?(\w+)`?\s*\((.*?);', 
        sql_content,
        re.DOTALL | re.IGNORECASE
    )

    # 解析每个表结构
    tables = []
    for stmt in statements:
        parsed = parse_table_statement(stmt)
        tables.append(parsed)
    
    return tables

def create_excel(tables, output_path):
    """生成带多sheet的Excel文件"""
    wb = Workbook()

    
    for table in tables:
        ws = wb.create_sheet(title=table["table_name"][:30])  # 限制sheet名称长度
        
        # 第一行写字段名
        headers = [field['name'] for field in table["fields"]]
        ws.append(headers)
        
        # 第二行写注释
        comments = [field["comment"] for field in table["fields"]]
        ws.append(comments)
        
        # 添加表注释说明（A1单元格）
        ws['A1'].comment = openpyxl.comments.Comment(f"表注释：{table['table_comment']}", "SQL Parser")
        
        
    if tables:
        del wb['Sheet']  # 删除默认sheet
    wb.save(output_path)

if __name__ == '__main__':
    # 使用示例
    tables = parse_sql_file(r'F:\test\ubuntu_configure\assist\python\tea_product\schema\medical_tab.sql')
    create_excel(tables, r'F:\test\ubuntu_configure\assist\python\tea_product\schema\database_structure.xlsx')