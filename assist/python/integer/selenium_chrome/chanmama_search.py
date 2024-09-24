from openpyxl import load_workbook
from openpyxl.styles import Font
from selenium.webdriver.common.by import By
from common.selenium_tools import SeleniumHelper
from pandas import DataFrame
from tqdm import tqdm
from openpyxl import load_workbook
import os
# 一.登录
# https://www.chanmama.com/awemeRank?keyword=&time=24h&digg=&gender type

domain= "https://www.chanmama.com"

def chanmama_search(url,driver_path,xls_path):
    """ 
    1.打开浏览器
    2.输入搜索关键词
    3.采集数据
    4.保存数据
    5.清洗数据
    """

    with SeleniumHelper(driver_path=driver_path) as driver:
    #热门视频页面
        #如果强制定向到登录界面
        if driver.wait_url_contains(url,"login"):
            #单击登录按钮
            driver.wait_presence_of_element(By.ID,"e2e-login-submit").click()()# 登录右侧 flex 对齐项目中心
            if not driver.wait_leave_of_element(By.XPATH,'//div[@class="login-r "]'):
                exit() #如果登录失败,结束程序

        #二.搜索
        key = input("请输入关键词:")
        driver.wait_presence_of_element(By.ID,'e2e-common-search-input').send_keys(key)
        driver.wait_presence_of_element(By.ID,'e2e-common-search-btn').click()
        
        #三.采集 
        rows = driver.wait_presence_of_elements(By.XPATH,'//tbody/tr[@data-v-450a04b2]')

        list_rows=[]
        for item in tqdm(rows):
            dict_row ={
            "视频标题":item.find_element(By.XPATH,'./td[1]//div[@class="ellipsis-2 text-align-left c333 pr5"]//a').text,
            "点赞数":item.find_element(By.XPATH,'./td[3]').text,
            "转发数":item.find_element(By.XPATH,'./td[5]').text,
            "视频链接":item.find_element(By.XPATH,'./td[6]//a').get_attribute("href")
            # "视频链接":domain+item.find_element(By.XPATH,'./td[6]//a').get_attribute("href")
            }
            list_rows.append(dict_row)
        file_name = os.path.join(xls_path, f"{key}热门短视频.xlsx")
        DataFrame(list_rows).to_excel(file_name,index=False)
        # 四.清洗
        # 加载Excel表格
        book = load_workbook(file_name)
        #找到第一张表格

        sheet = book.active

                
        for r in range(2,sheet.max_row+1): #注意:行列号从1开始
            for c in range(2,4):
                cell = sheet.cell(r,c)
                if type(cell.value)== str:
                    value = cell.value.replace(",","")
                    if value.endswith("w"):
                        cell.value =float(value[:-1])*10000
                    else:
                        cell.value = float(value)
                if cell.value >100000:
                #十六进制的颜色
                    cell.font=Font(color="ff0003")
        book.save(file_name)
    
if __name__=="__main__":
    url="https://www.douyin.com/user/MS4wLjABAAAAiWLBwkqSxfbwmsDwSiIgsMejChfAhO_U_FTyrqFFoio?from_tab_name=main"
    drive_path=r"D:\Tool\chromedriver-win64\chromedriver.exe"
    xls_path=r"F:\教程\三维讲解建筑结构图"
    chanmama_search(url,drive_path,xls_path)