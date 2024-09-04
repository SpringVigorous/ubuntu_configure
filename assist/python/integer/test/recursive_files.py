import os
from datetime import datetime
from openpyxl import Workbook

def list_files_and_folders(directory):
    """
    递归地列出目录及其子目录下的所有文件和文件夹，并返回文件和文件夹信息列表。
    """
    files_info = []
    folders_info = []

    for root, dirs, files in os.walk(directory):
        # 包括隐藏文件和文件夹
        files = [f for f in files if f[0] != '.'] + [f for f in files if f[0] == '.']
        dirs[:] = [d for d in dirs if d[0] != '.']  # 仅处理非隐藏目录
        
        # 处理文件夹信息
        for folder in dirs:
            folderpath = os.path.join(root, folder)
            creation_time = datetime.fromtimestamp(os.path.getctime(folderpath))
            modification_time = datetime.fromtimestamp(os.path.getmtime(folderpath))
            folders_info.append({
                '文件夹名': folder,
                '完整路径': folderpath,
                '创建日期': creation_time,
                '最后修改日期': modification_time
            })
        
        # 处理文件信息
        for file in files:
            filepath = os.path.join(root, file)
            filesize = os.path.getsize(filepath)
            creation_time = datetime.fromtimestamp(os.path.getctime(filepath))
            modification_time = datetime.fromtimestamp(os.path.getmtime(filepath))
            files_info.append({
                '文件名': file,
                '完整路径': filepath,
                '创建日期': creation_time,
                '文件大小': filesize,
                '最后修改日期': modification_time
            })

    return files_info, folders_info

def write_to_excel(files_info, folders_info, output_file):
    """
    将文件和文件夹信息写入 Excel 文件。
    """
    wb = Workbook()
    ws_files = wb.create_sheet("Files Information")
    ws_folders = wb.create_sheet("Folders Information")

    # 写入文件信息
    headers_files = ['文件名', '完整路径', '创建日期', '文件大小', '最后修改日期']
    ws_files.append(headers_files)
    for info in files_info:
        row = [
            info['文件名'],
            info['完整路径'],
            info['创建日期'].strftime('%Y-%m-%d %H:%M:%S'),
            info['文件大小'],
            info['最后修改日期'].strftime('%Y-%m-%d %H:%M:%S')
        ]
        ws_files.append(row)

    # 写入文件夹信息
    headers_folders = ['文件夹名', '完整路径', '创建日期', '最后修改日期']
    ws_folders.append(headers_folders)
    for info in folders_info:
        row = [
            info['文件夹名'],
            info['完整路径'],
            info['创建日期'].strftime('%Y-%m-%d %H:%M:%S'),
            info['最后修改日期'].strftime('%Y-%m-%d %H:%M:%S')
        ]
        ws_folders.append(row)

    # 保存 Excel 文件
    wb.save(output_file)

# 使用示例
directory_to_walk = r'C:\Users\Public'  # 注意反斜杠的使用
output_file = 'files_and_folders_info.xlsx'

# 获取文件和文件夹信息
files_info, folders_info = list_files_and_folders(directory_to_walk)

# 将文件和文件夹信息写入 Excel 文件
write_to_excel(files_info, folders_info, output_file)

print(f"文件和文件夹信息已成功输出到 {output_file}")