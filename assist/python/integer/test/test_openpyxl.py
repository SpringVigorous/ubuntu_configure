import openpyxl
import openpyxl.workbook

# wp=openpyxl.load_workbook(r"F:\cache\bilibili\263032155\599074\html\20240829\0_现代C++项目实战-tab.xlsx",read_only=False)
wp=openpyxl.Workbook()

print(wp.sheetnames)
titles=[sheet.title  for sheet in wp]
print(titles)
sheet1=wp.create_sheet(title="test", index=0)
del wp["Sheet"]
print(sheet1.title)
print(wp.sheetnames)